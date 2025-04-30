import os
import openpyxl
import csv

def xlsxToCsvHandler(xls_filename):
    # Extrai o nome do arquivo sem extensão e o diretório base.
    base_filename = os.path.splitext(os.path.basename(xls_filename))[0]
    base_dir = os.path.dirname(xls_filename)

    # Abre o workbook.
    try:
        wb = openpyxl.load_workbook(xls_filename)

        # Itera por todas as planilhas.
        for sheet_name in wb.sheetnames:
            try:
                # Abre a planilha pelo nome.
                wsh = wb[sheet_name]

                # Gera o nome do arquivo CSV.
                csv_filename = f"{base_dir}\\{base_filename}-{sheet_name}.csv"

                # Abre o arquivo CSV em modo de escrita com codificação utf-8.
                with open(csv_filename, "w", newline='', encoding='utf-8') as fh:
                    csv_out = csv.writer(fh)

                    # Itera pelas linhas da planilha e escreve no arquivo CSV.
                    for row in wsh.iter_rows():
                        csv_out.writerow([cell.value if cell.value is not None else '' for cell in row])

                print(f"Arquivo CSV {csv_filename} criado com sucesso.")
            except Exception as e:
                print(f"Erro ao criar o arquivo CSV para a planilha {sheet_name}.")
                print(e)
    except Exception as e:
        print("Erro ao abrir o arquivo Excel.")
        print(e)

if __name__ == '__main__':
    # Caminho completo para o novo arquivo Excel
    # xls_filename = r'C:\Users\murilo.ana\Downloads\ARQUIVO_EXEMPLO.xlsx'
    xls_filename = r'C:\Users\murilo.ana\Downloads\ARQUIVO_EXEMPLO.xlsx'
    xlsxToCsvHandler(xls_filename)
